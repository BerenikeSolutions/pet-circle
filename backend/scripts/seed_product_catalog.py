"""
PetCircle Phase 1 — Product Catalog Seed Script

Reads from PetCircle_Nutrition_Database_1.xlsx (4 data sheets)
and inserts products into the product_catalog table.

Usage:
    cd backend
    python -m scripts.seed_product_catalog

Requires: openpyxl, sqlalchemy, app.database configured
"""

import os
import sys
import logging
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

# Add backend to path for imports
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.database import SessionLocal
from app.models.product_catalog import ProductCatalog

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Path to Excel file (relative to project root)
EXCEL_PATH = Path(__file__).resolve().parent.parent.parent / "PetCircle_Nutrition_Database_1.xlsx"

# Cart item ID mapping from Notes column where "SKU in app: cX" appears
# Fallback manual mapping for known products
CART_ID_MAP = {
    "Drontal Plus": "c2",
    "Zesty Paws Salmon Oil": "c3",
    "Nutramax Cosequin DS": "c4",
    "NexGard": "c5",
    "Vitamin E-400 IU": "c6",
    "FortiFlora": "c11",
    "Calcitriol": "c12",
}


def safe_decimal(val, precision=2):
    """Convert a value to Decimal, returning None if invalid."""
    if val is None or val == "" or val == "-":
        return None
    try:
        return round(Decimal(str(val)), precision)
    except (InvalidOperation, ValueError):
        return None


def safe_int(val):
    """Convert a value to int, returning None if invalid."""
    if val is None or val == "" or val == "-":
        return None
    try:
        # Handle strings like "1,200" or "~500"
        cleaned = str(val).replace(",", "").replace("~", "").replace(">", "").replace("<", "").strip()
        # Handle ranges like "100-200" by taking the first value
        if "-" in cleaned and not cleaned.startswith("-"):
            cleaned = cleaned.split("-")[0].strip()
        return int(float(cleaned))
    except (ValueError, TypeError):
        return None


def safe_str(val, max_len=None):
    """Convert a value to string, replacing rupee symbols."""
    if val is None:
        return None
    s = str(val).replace("\u20b9", "Rs.").strip()
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s if s else None


def extract_cart_id(product_name, notes):
    """Extract cart_item_id from notes or product name matching."""
    if notes:
        notes_str = str(notes)
        # Look for "SKU in app: cX" pattern
        if "SKU" in notes_str and "c" in notes_str:
            import re
            match = re.search(r'c(\d+)', notes_str)
            if match:
                return f"c{match.group(1)}"

    # Fallback to name matching
    if product_name:
        for key, cart_id in CART_ID_MAP.items():
            if key.lower() in str(product_name).lower():
                return cart_id
    return None


def is_prescription(val):
    """Parse prescription field to boolean."""
    if val is None:
        return None
    s = str(val).lower().strip()
    if "yes" in s or "required" in s or "prescription" in s:
        return True
    if "no" in s or "otc" in s:
        return False
    return None


def seed_nutrition_database(wb, db):
    """Import food products from the Nutrition Database sheet."""
    ws = wb["Nutrition Database"]
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        brand = safe_str(row[0])
        product_name = safe_str(row[1])
        if not brand or not product_name:
            continue

        product = ProductCatalog(
            category="food",
            brand=brand,
            product_name=product_name,
            description=safe_str(row[2]),
            crude_protein=safe_decimal(row[3]),
            crude_fat=safe_decimal(row[4]),
            crude_fibre=safe_decimal(row[5]),
            moisture=safe_decimal(row[6]),
            ash=safe_decimal(row[7]),
            calcium=safe_decimal(row[8], 3),
            phosphorus=safe_decimal(row[9], 3),
            omega_3=safe_int(row[10]),
            omega_6=safe_int(row[11]),
            vitamin_e=safe_int(row[12]),
            vitamin_d3=safe_int(row[13]),
            glucosamine=safe_int(row[14]),
            probiotics=safe_str(row[15], 30),
            energy_kcal=safe_int(row[16]),
            life_stage=safe_str(row[17], 50),
            breed_size=safe_str(row[18], 50),
            type=safe_str(row[19], 50),
            pack_size=safe_str(row[20], 100),
            mrp=safe_str(row[21], 100),
            notes=safe_str(row[22]),
            cart_item_id=extract_cart_id(product_name, safe_str(row[22])),
        )
        db.add(product)
        count += 1

    logger.info("Nutrition Database: %d food products imported", count)
    return count


def seed_deworming(wb, db):
    """Import deworming medicines."""
    ws = wb["Deworming Medicines"]
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        brand = safe_str(row[0])
        product_name = safe_str(row[1])
        if not brand or not product_name:
            continue

        product = ProductCatalog(
            category="deworming",
            brand=brand,
            product_name=product_name,
            active_ingredient=safe_str(row[2]),
            indication=safe_str(row[3]),  # Mechanism / Targets
            dosage=safe_str(row[4]),
            frequency=safe_str(row[5], 200),
            formulation=safe_str(row[6], 50),
            prescription_required=is_prescription(row[7]),
            mrp=safe_str(row[8], 100),  # Pack / MRP
            notes=safe_str(row[9]),
            cart_item_id=extract_cart_id(product_name, safe_str(row[9])),
        )
        db.add(product)
        count += 1

    logger.info("Deworming Medicines: %d products imported", count)
    return count


def seed_flea_tick(wb, db):
    """Import flea & tick products."""
    ws = wb["Flea & Tick Products"]
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        brand = safe_str(row[0])
        product_name = safe_str(row[1])
        if not brand or not product_name:
            continue

        product = ProductCatalog(
            category="flea_tick",
            brand=brand,
            product_name=product_name,
            active_ingredient=safe_str(row[2]),
            indication=safe_str(row[3]),  # Targets / Coverage
            dosage=safe_str(row[4]),      # Dosage / Application
            frequency=safe_str(row[5], 200),  # Duration of Action
            formulation=safe_str(row[6], 50),
            prescription_required=is_prescription(row[7]),
            mrp=safe_str(row[8], 100),
            notes=safe_str(row[9]),
            cart_item_id=extract_cart_id(product_name, safe_str(row[9])),
        )
        db.add(product)
        count += 1

    logger.info("Flea & Tick Products: %d products imported", count)
    return count


def seed_disease_medicines(wb, db):
    """Import disease medicines."""
    ws = wb["Disease Medicines"]
    count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        # Column 0 is Disease/Condition Category, Column 1 is Medicine Name
        condition_category = safe_str(row[0])
        medicine_name = safe_str(row[1])
        if not medicine_name:
            continue

        # Use condition category as description, brand derived from medicine name
        brand = condition_category or "Various"

        product = ProductCatalog(
            category="medicine",
            brand=brand,
            product_name=medicine_name,
            active_ingredient=safe_str(row[2]),
            indication=safe_str(row[3]),
            dosage=safe_str(row[4]),
            frequency=safe_str(row[5], 200),
            formulation=safe_str(row[6], 50),
            prescription_required=is_prescription(row[7]),
            mrp=safe_str(row[8], 100),
            notes=safe_str(row[9]),
            cart_item_id=extract_cart_id(medicine_name, safe_str(row[9])),
        )
        db.add(product)
        count += 1

    logger.info("Disease Medicines: %d products imported", count)
    return count


def main():
    """Main entry point: load Excel and seed all product categories."""
    if not EXCEL_PATH.exists():
        logger.error("Excel file not found at %s", EXCEL_PATH)
        sys.exit(1)

    logger.info("Loading Excel file: %s", EXCEL_PATH)
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)

    db = SessionLocal()
    try:
        # Check if already seeded
        existing = db.query(ProductCatalog).count()
        if existing > 0:
            logger.warning("Product catalog already has %d rows. Skipping seed.", existing)
            logger.info("To re-seed, first run: DELETE FROM product_catalog;")
            return

        total = 0
        total += seed_nutrition_database(wb, db)
        total += seed_deworming(wb, db)
        total += seed_flea_tick(wb, db)
        total += seed_disease_medicines(wb, db)

        db.commit()
        logger.info("Total products seeded: %d", total)

    except Exception as e:
        db.rollback()
        logger.error("Seed failed: %s", e)
        raise
    finally:
        db.close()
        wb.close()


if __name__ == "__main__":
    main()
